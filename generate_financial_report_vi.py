#!/usr/bin/env python3
"""
Tạo báo cáo tài chính chi tiết dạng Excel (100% Tiếng Việt)
"""

import sqlite3
import os
from datetime import datetime, timedelta
from pathlib import Path
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(["pip3", "install", "openpyxl"], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

DB_PATH = os.path.expanduser("~/trighub_transactions.db")
REPORT_DIR = os.path.expanduser("~/Financial_Reports")

class FinancialReportGenerator:
    """Tạo báo cáo tài chính Excel"""
    
    def __init__(self):
        self.db_path = DB_PATH
        Path(REPORT_DIR).mkdir(exist_ok=True)
        
        # Định dạng
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.category_font = Font(bold=True, size=11)
        self.total_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.total_font = Font(bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _get_data(self, days=30):
        """Lấy dữ liệu giao dịch"""
        conn = sqlite3.connect(self.db_path)
        
        cutoff_date = (datetime.now() - timedelta(days=days)).isoformat()
        
        # Tóm tắt
        income = conn.execute(
            'SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE transaction_type="IN" AND timestamp >= ?',
            (cutoff_date,)
        ).fetchone()[0]
        
        expense = conn.execute(
            'SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE transaction_type="OUT" AND timestamp >= ?',
            (cutoff_date,)
        ).fetchone()[0]
        
        # Theo danh mục
        categories = conn.execute('''
            SELECT category, transaction_type, COUNT(*) as count, SUM(amount) as total
            FROM transactions
            WHERE timestamp >= ?
            GROUP BY category, transaction_type
            ORDER BY total DESC
        ''', (cutoff_date,)).fetchall()
        
        # Hàng ngày
        daily = conn.execute('''
            SELECT DATE(timestamp),
                   SUM(CASE WHEN transaction_type='IN' THEN amount ELSE 0 END) as income,
                   SUM(CASE WHEN transaction_type='OUT' THEN amount ELSE 0 END) as expense
            FROM transactions
            WHERE timestamp >= ?
            GROUP BY DATE(timestamp)
            ORDER BY DATE(timestamp)
        ''', (cutoff_date,)).fetchall()
        
        # Top giao dịch
        top_txns = conn.execute('''
            SELECT timestamp, amount, content, bank_name, transaction_type
            FROM transactions
            WHERE timestamp >= ?
            ORDER BY amount DESC
            LIMIT 10
        ''', (cutoff_date,)).fetchall()
        
        conn.close()
        
        return {
            'income': income,
            'expense': expense,
            'net': income - expense,
            'categories': categories,
            'daily': daily,
            'top_txns': top_txns,
            'days': days
        }
    
    def _format_type(self, txn_type):
        """Chuyển IN/OUT sang Tiếng Việt"""
        return "Thu" if txn_type == "IN" else "Chi"
    
    def _add_summary_sheet(self, wb, data):
        """Thêm sheet Tóm tắt"""
        ws = wb.active
        ws.title = "Tóm tắt"
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        row = 1
        
        # Tiêu đề
        ws[f'A{row}'] = "BÁO CÁO TÀI CHÍNH"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = f"Kỳ: {data['days']} ngày gần nhất"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        row += 1
        
        ws[f'A{row}'] = f"Tạo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        row += 2
        
        # Bảng tóm tắt
        ws[f'A{row}'] = "TÓM TẮT"
        ws[f'A{row}'].font = self.category_font
        ws[f'A{row}'].fill = self.category_fill
        row += 1
        
        summary_items = [
            ("💰 Tổng Thu nhập", data['income'], "IN"),
            ("💸 Tổng Chi phí", data['expense'], "OUT"),
            ("💵 Lưu lại ròng", data['net'], "NET"),
        ]
        
        for label, value, _type in summary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '#,##0'
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            
            if _type == "IN":
                ws[f'B{row}'].font = Font(color="008000", bold=True)
            elif _type == "OUT":
                ws[f'B{row}'].font = Font(color="FF0000", bold=True)
            else:
                ws[f'B{row}'].fill = self.total_fill
                ws[f'B{row}'].font = self.total_font
            
            row += 1
        
        # Tỷ lệ
        row += 1
        savings_rate = (data['net'] / data['income'] * 100) if data['income'] > 0 else 0
        
        ws[f'A{row}'] = "📊 Tỷ lệ tiết kiệm"
        ws[f'B{row}'] = f"{savings_rate:.1f}%"
        row += 1
        
        expense_ratio = (data['expense'] / data['income'] * 100) if data['income'] > 0 else 0
        ws[f'A{row}'] = "📈 Tỷ lệ chi phí"
        ws[f'B{row}'] = f"{expense_ratio:.1f}%"
    
    def _add_category_sheet(self, wb, data):
        """Thêm sheet Danh mục"""
        ws = wb.create_sheet("Danh mục")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        
        row = 1
        
        # Header
        headers = ["📂 Danh mục", "🏷️ Loại", "🔢 Số lần", "💰 Số tiền"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        row += 1
        
        # Dữ liệu
        for category, transaction_type, count, amount in data['categories']:
            type_label = self._format_type(transaction_type)
            ws.cell(row=row, column=1, value=f"{category} ({type_label})")
            ws.cell(row=row, column=2, value=type_label)
            ws.cell(row=row, column=3, value=count)
            ws.cell(row=row, column=4, value=amount)
            
            ws.cell(row=row, column=4).number_format = '#,##0'
            
            # Màu sắc
            if transaction_type == "IN":
                ws.cell(row=row, column=2).font = Font(color="008000")
            else:
                ws.cell(row=row, column=2).font = Font(color="FF0000")
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
    
    def _add_daily_sheet(self, wb, data):
        """Thêm sheet Hàng ngày"""
        ws = wb.create_sheet("Hàng ngày")
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        row = 1
        
        # Header
        headers = ["📅 Ngày", "📥 Thu nhập", "📤 Chi phí", "💵 Lưu lại"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        row += 1
        
        # Dữ liệu
        for date, income, expense in data['daily']:
            income = income or 0
            expense = expense or 0
            net = income - expense
            
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=income)
            ws.cell(row=row, column=3, value=expense)
            ws.cell(row=row, column=4, value=net)
            
            for col in range(2, 5):
                ws.cell(row=row, column=col).number_format = '#,##0'
            
            # Màu net
            if net >= 0:
                ws.cell(row=row, column=4).font = Font(color="008000")
            else:
                ws.cell(row=row, column=4).font = Font(color="FF0000")
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
    
    def _add_transactions_sheet(self, wb, data):
        """Thêm sheet Top Giao dịch"""
        ws = wb.create_sheet("Top Giao dịch")
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 8
        
        row = 1
        
        # Header
        headers = ["📅 Ngày", "💰 Số tiền", "📝 Mô tả", "🏦 Ngân hàng", "🏷️"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        row += 1
        
        # Dữ liệu
        for timestamp, amount, content, bank, transaction_type in data['top_txns']:
            type_label = self._format_type(transaction_type)
            ws.cell(row=row, column=1, value=timestamp[:10])
            ws.cell(row=row, column=2, value=amount)
            ws.cell(row=row, column=3, value=content[:50])
            ws.cell(row=row, column=4, value=bank)
            ws.cell(row=row, column=5, value=type_label)
            
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            # Màu type
            if transaction_type == "IN":
                ws.cell(row=row, column=5).font = Font(color="008000", bold=True)
            else:
                ws.cell(row=row, column=5).font = Font(color="FF0000", bold=True)
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
    
    def generate(self, days=30, filename=None):
        """Tạo báo cáo hoàn chỉnh"""
        if filename is None:
            month = datetime.now().strftime("%Y-%m")
            filename = f"Báo_cáo_tài_chính_{month}.xlsx"
        
        filepath = os.path.join(REPORT_DIR, filename)
        
        print(f"📊 Tạo báo cáo tài chính cho {days} ngày gần nhất...")
        
        # Lấy dữ liệu
        data = self._get_data(days)
        
        # Tạo workbook
        wb = Workbook()
        
        # Thêm sheets
        self._add_summary_sheet(wb, data)
        self._add_category_sheet(wb, data)
        self._add_daily_sheet(wb, data)
        self._add_transactions_sheet(wb, data)
        
        # Lưu
        wb.save(filepath)
        
        print(f"✅ Báo cáo đã lưu: {filepath}")
        print(f"   📥 Thu nhập:  {data['income']:,} VND")
        print(f"   📤 Chi phí:   {data['expense']:,} VND")
        print(f"   💵 Lưu lại:   {data['net']:,} VND")
        
        return filepath

def main():
    """Điểm vào chính"""
    import sys
    
    generator = FinancialReportGenerator()
    
    days = 30
    filename = None
    
    if len(sys.argv) > 1:
        days = int(sys.argv[1])
    
    if len(sys.argv) > 2:
        filename = sys.argv[2]
    
    filepath = generator.generate(days=days, filename=filename)
    print(f"\n📁 Vị trí file: {filepath}")

if __name__ == "__main__":
    main()
